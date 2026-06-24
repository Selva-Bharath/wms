from flask import Blueprint, request, jsonify
from models.database import db
from models.attendance import Attendance
from datetime import datetime
from models.employee import Employee
from models.user import User
from datetime import date
from sqlalchemy import extract
from datetime import timedelta
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from flask import send_file
from models.leave import LeaveRequest, LeaveLedger
from io import BytesIO


from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Border,
    Side,
    Alignment
)
from openpyxl.utils import get_column_letter


attendance_bp = Blueprint(
    "attendance",
    __name__
)


@attendance_bp.route("/checkin", methods=["POST"])
def check_in():

    try:

        data = request.json

        user_id = data.get("user_id")

        if not user_id:
            return jsonify({
                "success": False,
                "message": "User ID is required"
            }), 400

        employee = Employee.query.filter_by(
            user_id=user_id
        ).first()

        if not employee:
            return jsonify({
                "success": False,
                "message": "Employee not found"
            }), 404

        # =====================================
        # SHIFT VALIDATION
        # =====================================

        shift_name = (
            employee.shift_timing or ""
        ).strip().lower()

        current_time = datetime.now().time()

        print("================================")
        print("Employee Shift:", employee.shift_timing)
        print("Current Time:", current_time)
        print("================================")

        # First Shift (06:00 AM - 02:00 PM)
        if shift_name == "first shift":
            print("================================")
            print("USER ID:", user_id)
            print("EMPLOYEE:", employee.first_name)
            print("SHIFT RAW:", employee.shift_timing)
            print("SHIFT LOWER:", shift_name)
            print("CURRENT TIME:", current_time)
            print("================================")

            allowed_time = datetime.strptime(
                "06:00",
                "%H:%M"
            ).time()

            if current_time < allowed_time:
                return jsonify({
                    "success": False,
                    "message": "First Shift check-in allowed only after 06:00 AM"
                }), 400

        # General Shift (09:00 AM - 06:00 PM)
        elif shift_name == "general shift":

            allowed_time = datetime.strptime(
                "09:00",
                "%H:%M"
            ).time()

            if current_time < allowed_time:
                return jsonify({
                    "success": False,
                    "message": "General Shift check-in allowed only after 09:00 AM"
                }), 400

        # Second Shift (02:00 PM - 10:00 PM)
        elif shift_name == "second shift":

            allowed_time = datetime.strptime(
                "14:00",
                "%H:%M"
            ).time()

            if current_time < allowed_time:
                return jsonify({
                    "success": False,
                    "message": "Second Shift check-in allowed only after 02:00 PM"
                }), 400

        # Night Shift (10:00 PM - 06:00 AM)
        elif shift_name == "night shift":

            allowed_time = datetime.strptime(
                "22:00",
                "%H:%M"
            ).time()

            if current_time < allowed_time:
                return jsonify({
                    "success": False,
                    "message": "Night Shift check-in allowed only after 10:00 PM"
                }), 400

        # =====================================
        # CHECK ALREADY CHECKED IN
        # =====================================

        today = datetime.now().date()

        attendance = Attendance.query.filter_by(
            user_id=user_id,
            attendance_date=today
        ).first()

        if attendance:
            return jsonify({
                "success": False,
                "message": "You have already checked in today."
            }), 400

        # =====================================
        # CREATE ATTENDANCE
        # =====================================

        attendance = Attendance(
            user_id=user_id,
            attendance_date=today,
            check_in=datetime.now(),
            status="Present"
        )

        db.session.add(attendance)
        db.session.commit()

        return jsonify({
            "success": True,
            "message": "Checked In Successfully"
        })

    except Exception as e:

        db.session.rollback()

        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


@attendance_bp.route("/checkout", methods=["POST"])
def check_out():

    try:

        data = request.json

        user_id = data.get("user_id")

        if not user_id:
            return jsonify({
                "success": False,
                "error": "User ID is required"
            }), 400

        print("CHECKOUT USER ID:", user_id)

        attendance = Attendance.query.filter(
            Attendance.user_id == user_id,
            Attendance.check_in.isnot(None),
            Attendance.check_out.is_(None)
        ).order_by(
            Attendance.id.desc()
        ).first()

        print("ATTENDANCE FOUND:", attendance)

        if not attendance:
            return jsonify({
                "success": False,
                "error": "No active check-in found"
            }), 404

        if not attendance.check_in:
            return jsonify({
                "success": False,
                "error": "Check-in time missing"
            }), 400

        attendance.check_out = datetime.now()

        total_seconds = (
            attendance.check_out -
            attendance.check_in
        ).total_seconds()

        break_minutes = (
            attendance.total_break_minutes or 0
        )

        total_seconds -= break_minutes * 60

        attendance.total_hours = round(
            total_seconds / 3600,
            2
        )

        attendance.status = "Present"

        db.session.commit()

        return jsonify({
            "success": True,
            "message": "Checked Out Successfully",
            "check_in": attendance.check_in.strftime("%Y-%m-%d %H:%M:%S"),
            "check_out": attendance.check_out.strftime("%Y-%m-%d %H:%M:%S"),
            "total_hours": attendance.total_hours
        }), 200

    except Exception as e:

        db.session.rollback()

        print("CHECKOUT ERROR:", str(e))

        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


@attendance_bp.route("/status/<int:user_id>")
def attendance_status(user_id):

    attendance = Attendance.query.filter_by(
        user_id=user_id,
        check_out=None
    ).order_by(
        Attendance.id.desc()
    ).first()

    if not attendance:
        return jsonify({
            "checked_in": False
        })

    return jsonify({
        "checked_in": True,
        "check_in": attendance.check_in.isoformat(),
        "lunch_break": attendance.lunch_break,
        "tea_break": attendance.tea_break,
        "lunch_start": attendance.lunch_start.isoformat() if attendance.lunch_start else None,
        "tea_start": attendance.tea_start.isoformat() if attendance.tea_start else None,
        "lunch_minutes": attendance.lunch_minutes or 0,
        "tea_minutes": attendance.tea_minutes or 0,
        "total_break_minutes": attendance.total_break_minutes or 0
    })


@attendance_bp.route("/lunch-break", methods=["POST"])
def lunch_break():

    try:

        data = request.json

        attendance = Attendance.query.filter_by(
            user_id=data["user_id"],
            check_out=None
        ).order_by(
            Attendance.id.desc()
        ).first()

        if not attendance:
            return jsonify({
                "success": False,
                "error": "Attendance not found"
            }), 404

        action = data.get("action")

        if action == "start":
            attendance.lunch_break = True
            attendance.lunch_start = datetime.now()

        elif action == "stop":
            attendance.lunch_break = False
            attendance.lunch_end = datetime.now()

        if attendance.lunch_start and attendance.lunch_end:
            attendance.lunch_minutes = int(
                (
                    attendance.lunch_end -
                    attendance.lunch_start
                ).total_seconds() / 60
            )

        attendance.total_break_minutes = (
            (attendance.lunch_minutes or 0) +
            (attendance.tea_minutes or 0)
        )

        db.session.commit()

        return jsonify({"success": True})

    except Exception as e:

        print("LUNCH BREAK ERROR:", str(e))

        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


@attendance_bp.route("/tea-break", methods=["POST"])
def tea_break():

    try:

        data = request.json

        attendance = Attendance.query.filter_by(
            user_id=data["user_id"],
            check_out=None
        ).order_by(
            Attendance.id.desc()
        ).first()

        if not attendance:
            return jsonify({
                "success": False,
                "error": "Attendance not found"
            }), 404

        action = data.get("action")

        if action == "start":
            attendance.tea_break = True
            attendance.tea_start = datetime.now()

        elif action == "stop":
            attendance.tea_break = False
            attendance.tea_end = datetime.now()

        if attendance.tea_start and attendance.tea_end:
            attendance.tea_minutes = int(
                (
                    attendance.tea_end -
                    attendance.tea_start
                ).total_seconds() / 60
            )

        attendance.total_break_minutes = (
            (attendance.lunch_minutes or 0) +
            (attendance.tea_minutes or 0)
        )

        db.session.commit()

        return jsonify({"success": True})

    except Exception as e:

        print("TEA BREAK ERROR:", str(e))

        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


@attendance_bp.route("/history/<int:user_id>")
def attendance_history(user_id):

    records = Attendance.query.filter_by(
        user_id=user_id
    ).order_by(
        Attendance.id.desc()
    ).all()

    result = []

    for record in records:
        result.append({
            "id": record.id,
            "date": record.attendance_date.strftime("%Y-%m-%d") if record.attendance_date else "-",
            "checkIn": record.check_in.strftime("%I:%M %p") if record.check_in else "-",
            "checkOut": record.check_out.strftime("%I:%M %p") if record.check_out else "-",
            "workingHours": record.total_hours,
            "lunchMinutes": record.lunch_minutes,
            "teaMinutes": record.tea_minutes,
            "totalBreak": record.total_break_minutes,
            "status": record.status
        })

    return jsonify(result)


@attendance_bp.route("/", methods=["GET"])
def get_attendance():

    today = datetime.now().date()

    employees = Employee.query.all()

    attendance_list = []

    for employee in employees:

        attendance = Attendance.query.filter_by(
            user_id=employee.user_id,
            attendance_date=today
        ).first()

        if attendance:
            status = attendance.status or "Present"
            check_in = attendance.check_in.strftime("%H:%M:%S") if attendance.check_in else "-"
            check_out = attendance.check_out.strftime("%H:%M:%S") if attendance.check_out else "-"
            total_hours = attendance.total_hours
        else:
            status = "Absent"
            check_in = "-"
            check_out = "-"
            total_hours = 0

        attendance_list.append({
            "user_id": employee.user_id,
            "employee_name": f"{employee.first_name} {employee.last_name}",
            "department": employee.department,
            "designation": employee.designation,
            "check_in": check_in,
            "check_out": check_out,
            "total_hours": total_hours,
            "attendance_date": str(today),
            "status": status,
            "shift_timing": (
                attendance.shift_timing
                if attendance and attendance.shift_timing
                else employee.shift_timing or "General Shift"
            )
        })

    return jsonify(attendance_list)


@attendance_bp.route("/generate-daily-attendance")
def generate_daily_attendance():

    today = date.today()

    users = User.query.filter_by(
        is_active=True
    ).all()

    count = 0

    for user in users:

        existing = Attendance.query.filter_by(
            user_id=user.id,
            attendance_date=today
        ).first()

        if not existing:
            attendance = Attendance(
                user_id=user.id,
                attendance_date=today,
                status="Absent"
            )
            db.session.add(attendance)
            count += 1

    db.session.commit()

    return jsonify({
        "success": True,
        "records_created": count
    })


@attendance_bp.route("/weekly", methods=["GET"])
def get_weekly_attendance():

    try:

        result = []

        employees = Employee.query.all()

        for i in range(7):

            current_date = date.today() - timedelta(days=i)

            for employee in employees:

                attendance = Attendance.query.filter_by(
                    user_id=employee.user_id,
                    attendance_date=current_date
                ).first()

                result.append({
                    "employee_name": f"{employee.first_name} {employee.last_name}",
                    "team": employee.department if employee.department else "-",
                    "date": current_date.strftime("%d-%m-%Y"),
                    "check_in": attendance.check_in.strftime("%I:%M %p") if attendance and attendance.check_in else "-",
                    "check_out": attendance.check_out.strftime("%I:%M %p") if attendance and attendance.check_out else "-",
                    "total_hours": attendance.total_hours if attendance else "-",
                    "status": attendance.status if attendance else "Absent",
                    "shift_timing": (
                        attendance.shift_timing
                        if attendance and attendance.shift_timing
                        else employee.shift_timing or "General Shift"
                    )
                })

        return jsonify(result)

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@attendance_bp.route("/monthly", methods=["GET"])
def get_monthly_attendance():

    try:

        result = []

        employees = Employee.query.all()

        for i in range(30):

            current_date = date.today() - timedelta(days=i)

            for employee in employees:

                attendance = Attendance.query.filter_by(
                    user_id=employee.user_id,
                    attendance_date=current_date
                ).first()

                result.append({
                    "employee_name": f"{employee.first_name} {employee.last_name}",
                    "team": employee.department if employee.department else "-",
                    "date": current_date.strftime("%d-%m-%Y"),
                    "check_in": attendance.check_in.strftime("%I:%M %p") if attendance and attendance.check_in else "-",
                    "check_out": attendance.check_out.strftime("%I:%M %p") if attendance and attendance.check_out else "-",
                    "total_hours": attendance.total_hours if attendance else "-",
                    "status": attendance.status if attendance else "Absent",
                    "shift_timing": (
                        attendance.shift_timing
                        if attendance and attendance.shift_timing
                        else employee.shift_timing or "General Shift"
                    )
                })

        return jsonify(result)

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@attendance_bp.route("/export-monthly", methods=["GET"])
def export_monthly_attendance():

    try:

        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance Report"

        today = date.today()

        if today.day < 25:
            if today.month == 1:
                start_date = date(today.year - 1, 11, 25)
                end_date = date(today.year - 1, 12, 24)
            else:
                start_date = date(today.year, today.month - 2, 25)
                end_date = date(today.year, today.month - 1, 24)
        else:
            if today.month == 1:
                start_date = date(today.year - 1, 12, 25)
                end_date = date(today.year, 1, 24)
            else:
                start_date = date(today.year, today.month - 1, 25)
                end_date = date(today.year, today.month, 24)

        # =====================================
        # STYLES
        # =====================================

        purple_fill = PatternFill(fill_type="solid", fgColor="B58CE5")
        yellow_fill = PatternFill(fill_type="solid", fgColor="F7F1A0")
        white_font = Font(bold=True, color="FFFFFF", size=12)
        bold_font = Font(bold=True, size=12)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        # =====================================
        # TITLE
        # =====================================

        ws.merge_cells("A1:K1")
        ws["A1"] = "ATTENDANCE REPORT"
        ws["A1"].fill = purple_fill
        ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

        # =====================================
        # MONTH HEADER
        # =====================================

        ws.merge_cells("A2:K2")
        ws["A2"] = f"Attendance Summary {date.today().strftime('%B %Y')}"
        ws["A2"].fill = purple_fill
        ws["A2"].font = white_font
        ws["A2"].alignment = Alignment(horizontal="center")

        # =====================================
        # DATE RANGE
        # =====================================

        ws.merge_cells("A3:K3")
        ws["A3"] = (
            f"Attendance Cycle : "
            f"{start_date.strftime('%d-%b-%Y')} "
            f"to "
            f"{end_date.strftime('%d-%b-%Y')}"
        )
        ws["A3"].fill = yellow_fill
        ws["A3"].font = bold_font
        ws["A3"].alignment = Alignment(horizontal="center")

        # =====================================
        # COLUMN HEADERS
        # =====================================

        headers = [
            "S.No", "Emp Code", "Emp Name", "D.O.J", "Department",
            "Total Days In Cycle", "Present Days", "Approved Leave Days",
            "Absent Days", "Date Of Leave", "Remarks"
        ]

        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=5, column=col_num)
            cell.value = header
            cell.fill = purple_fill
            cell.font = white_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # =====================================
        # EMPLOYEE DATA
        # =====================================

        employees = Employee.query.all()
        row = 6

        for index, employee in enumerate(employees, start=1):

            attendance_records = Attendance.query.filter(
                Attendance.user_id == employee.user_id,
                Attendance.attendance_date >= start_date,
                Attendance.attendance_date <= end_date
            ).all()

            leave_requests = LeaveRequest.query.filter(
                LeaveRequest.employee_id == employee.id,
                LeaveRequest.status == "Approved"
            ).all()

            days_worked = len([a for a in attendance_records if a.status == "Present"])
            total_days_cycle = (end_date - start_date).days + 1

            total_leaves = 0
            leave_dates_list = []

            for leave in leave_requests:
                total_leaves += (leave.total_days or 0)
                if leave.from_date and leave.to_date:
                    leave_dates_list.append(
                        f"{leave.from_date.strftime('%d-%b-%Y')} "
                        f"to "
                        f"{leave.to_date.strftime('%d-%b-%Y')}"
                    )

            approved_leave_days = total_leaves

            absent_days = total_days_cycle - days_worked - approved_leave_days
            if absent_days < 0:
                absent_days = 0

            leave_dates = ", ".join(leave_dates_list)
            remarks = ""

            ws.cell(row=row, column=1).value = index
            ws.cell(row=row, column=2).value = (
                employee.employee_id if hasattr(employee, "employee_id") else employee.user_id
            )
            ws.cell(row=row, column=3).value = f"{employee.first_name} {employee.last_name}"
            ws.cell(row=row, column=4).value = (
                str(employee.joining_date)
                if hasattr(employee, "joining_date") and employee.joining_date
                else ""
            )
            ws.cell(row=row, column=5).value = employee.department if employee.department else "-"
            ws.cell(row=row, column=6).value = total_days_cycle
            ws.cell(row=row, column=7).value = days_worked
            ws.cell(row=row, column=8).value = approved_leave_days
            ws.cell(row=row, column=9).value = absent_days
            ws.cell(row=row, column=10).value = leave_dates
            ws.cell(row=row, column=11).value = remarks

            for col in range(1, 12):

                cell = ws.cell(row=row, column=col)

                cell.border = thin_border

    # Number columns center
                if col in [1, 6, 7, 8, 9]:
                    cell.alignment = Alignment(
                        horizontal="center",
                        vertical="center"
                    )

    # Text columns left
                else:
                    cell.alignment = Alignment(
                    horizontal="left",
                    vertical="center"
                )

            row += 1

        # =====================================
        # AUTO WIDTH
        # =====================================

        for column_cells in ws.columns:
            length = max(
                len(str(cell.value)) if cell.value else 0
                for cell in column_cells
            )
            ws.column_dimensions[
                get_column_letter(column_cells[0].column)
            ].width = length + 5

        ws.column_dimensions["A"].width = 6

        # =====================================
        # FILTER
        # =====================================

        ws.auto_filter.ref = f"A5:K{row}"

        # =====================================
        # SAVE FILE
        # =====================================

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name="Attendance_Report.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


@attendance_bp.route("/credit-monthly-leaves", methods=["POST"])
def credit_monthly_leaves():

    try:

        current_month = datetime.now().strftime("%B")
        current_year = datetime.now().year

        employees = Employee.query.all()

        for employee in employees:

            existing = LeaveLedger.query.filter_by(
                employee_id=employee.employee_id,
                month=current_month,
                year=current_year
            ).first()

            if existing:
                continue

            opening_cl = employee.casual_leave or 0
            opening_sl = employee.sick_leave or 0
            opening_el = employee.earned_leave or 0

            credit_cl = 5
            credit_sl = 5
            credit_el = 5

            closing_cl = opening_cl + credit_cl
            closing_sl = opening_sl + credit_sl
            closing_el = opening_el + credit_el

            employee.casual_leave = closing_cl
            employee.sick_leave = closing_sl
            employee.earned_leave = closing_el

            ledger = LeaveLedger(
                employee_id=employee.employee_id,
                month=current_month,
                year=current_year,
                opening_cl=opening_cl,
                opening_sl=opening_sl,
                opening_el=opening_el,
                credit_cl=credit_cl,
                credit_sl=credit_sl,
                credit_el=credit_el,
                closing_cl=closing_cl,
                closing_sl=closing_sl,
                closing_el=closing_el
            )

            db.session.add(ledger)

        db.session.commit()

        return jsonify({
            "success": True,
            "message": "Leave credited successfully"
        })

    except Exception as e:

        db.session.rollback()

        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


@attendance_bp.route("/export-paysheet", methods=["GET"])
def export_paysheet():

    try:

        wb = Workbook()
        ws = wb.active
        ws.title = "Paysheet"

        today = date.today()

        if today.day < 25:
            if today.month == 1:
                start_date = date(today.year - 1, 11, 25)
                end_date = date(today.year - 1, 12, 24)
            else:
                start_date = date(today.year, today.month - 2, 25)
                end_date = date(today.year, today.month - 1, 24)
        else:
            if today.month == 1:
                start_date = date(today.year - 1, 12, 25)
                end_date = date(today.year, 1, 24)
            else:
                start_date = date(today.year, today.month - 1, 25)
                end_date = date(today.year, today.month, 24)

        header_fill = PatternFill(fill_type="solid", fgColor="D9A066")
        header_font = Font(bold=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        # =====================================
        # TITLE ROWS
        # =====================================

        # FIX 1: these 3 blocks were at wrong indentation (module level)
        # corrected to be inside the try block
        ws.merge_cells("A1:BE1")
        ws["A1"] = "S4 CARLISLE PUBLISHING SERVICES"
        ws["A1"].font = Font(bold=True, size=18)
        ws["A1"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A2:BE2")
        ws["A2"] = "MONTHLY PAYSHEET REPORT"
        ws["A2"].font = Font(bold=True, size=14)
        ws["A2"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A3:BE3")
        ws["A3"] = (
            f"Payroll Cycle : "
            f"{start_date.strftime('%d-%b-%Y')} "
            f"to "
            f"{end_date.strftime('%d-%b-%Y')}"
        )
        ws["A3"].font = Font(bold=True, size=12)
        ws["A3"].alignment = Alignment(horizontal="center")

        # =====================================
        # COLUMN HEADERS
        # =====================================

        headers = [
            "S.No", "EMP NO", "Gender", "PF No", "UAN No", "ESI No",
            "Employee Name", "Department", "Designation", "Mail ID", "DOJ",
            "No Of Days In Month", "Days Payable",
            "Basic", "HRA", "LTA", "Other Allowance", "Gross Salary",
            "Earned Basic", "Earned HRA", "Earned LTA", "Earned Other Allowance", "Earned Actual Gross",
            "Attendance Bonus", "ODW", "Total",
            "Internet Charges",
            "Gross Earned Salary", "Earned PF Wages",
            "PF Ded Employee", "PF Ded Employer",
            "VPF",
            "PF & VPF Ded Employee",
            "ESI Ded Employee", "ESI Ded Employer",
            "Salary Advance",
            "TDS", "LWF", "PT",
            "Other Deduction",
            "Total Deduction",
            "Net Transfer",
            "Account No", "IFSC Code", "Branch Code",
            "PF Wage", "PF",
            "EPS Wage",
            "8.33 %", "3.67 %", "0.50 %", "0.50 % Employer", "0.01 %",
            "Bonus",
            "Actual Month CTC", "Earned Month CTC",
            "Remarks"
        ]

        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=5, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border

        employees = Employee.query.all()
        row = 6

        for index, employee in enumerate(employees, start=1):

            attendance_records = Attendance.query.filter(
                Attendance.user_id == employee.user_id,
                Attendance.attendance_date >= start_date,
                Attendance.attendance_date <= end_date
            ).all()

            leave_requests = LeaveRequest.query.filter(
                LeaveRequest.employee_id == str(employee.id),
                LeaveRequest.status == "Approved"
            ).all()

            total_leaves = sum(leave.total_days or 0 for leave in leave_requests)
            total_days_cycle = (end_date - start_date).days + 1

            present_days = len([a for a in attendance_records if a.status == "Present"])

            days_payable = present_days + total_leaves

            if days_payable > total_days_cycle:
                days_payable = total_days_cycle

            absent_days = total_days_cycle - present_days - total_leaves
            if absent_days < 0:
                absent_days = 0

            salary = employee.salary or 0

            hra = 0
            lta = 0
            other_allowance = 0
            pf_deduction = 0
            esi_deduction = 0
            tds = 0
            pt = 0
            lwf = 0
            bonus = 0
            internet_charges = 0
            salary_advance = 0

            actual_ctc = salary

            earned_ctc = round(
                (salary / total_days_cycle) * days_payable, 2
            ) if total_days_cycle > 0 else 0

            net_transfer = round(
                earned_ctc - (
                    pf_deduction +
                    esi_deduction +
                    tds +
                    pt +
                    lwf +
                    salary_advance
                ),
                2
            )

            data = [
                index,
                employee.employee_id,
                employee.gender,
                employee.pf_number,
                employee.uan_number,
                employee.esi_number,
                f"{employee.first_name} {employee.last_name}",
                employee.department,
                employee.designation,
                employee.email,
                str(employee.joining_date),
                total_days_cycle,
                days_payable,
                salary,
                hra,
                lta,
                other_allowance,
                salary,
                "",  # Earned Basic
                "",  # Earned HRA
                "",  # Earned LTA
                "",  # Earned Other Allowance
                "",  # Earned Actual Gross
                "",  # Attendance Bonus
                "",  # ODW
                "",  # Total
                internet_charges,
                earned_ctc,
                "",  # Earned PF Wages
                pf_deduction,
                "",  # PF Ded Employer
                "",  # VPF
                "",  # PF & VPF Ded Employee
                esi_deduction,
                "",  # ESI Ded Employer
                salary_advance,
                tds,
                lwf,
                pt,
                "",  # Other Deduction
                "",  # Total Deduction
                net_transfer,
                employee.account_number,
                employee.ifsc_code,
                "",  # Branch Code
                "",  # PF Wage
                "",  # PF
                "",  # EPS Wage
                "",  # 8.33%
                "",  # 3.67%
                "",  # 0.50%
                "",  # 0.50% Employer
                "",  # 0.01%
                bonus,
                actual_ctc,
                earned_ctc,
                ""   # Remarks
            ]

            # FIX 2: removed the dangling if/else block that referenced
            # cell before it was assigned; alignment is now handled cleanly
            # after cell is created below
            for col_num, value in enumerate(data, start=1):
                cell = ws.cell(row=row, column=col_num, value=value)
                cell.border = thin_border

                if col_num == 1:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif col_num in [
                    2, 3, 4, 5, 6,
                    11, 12, 13, 14, 15, 16,
                    17, 18, 19, 20, 21, 22,
                    23, 24, 25, 26, 27, 28,
                    29, 30, 31, 32, 33, 34,
                    35, 36, 37, 38, 39, 40,
                    41, 42, 43, 44, 45, 46,
                    47, 48, 49, 50, 51, 52,
                    53, 54, 55, 56, 57, 58,
                    59
                ]:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

            row += 1

        # =====================================
        # AUTO WIDTH
        # =====================================

        for column_cells in ws.columns:
            try:
                length = max(
                    len(str(cell.value)) if cell.value else 0
                    for cell in column_cells
                )
                ws.column_dimensions[
                    get_column_letter(column_cells[0].column)
                ].width = length + 5
            except Exception:
                pass

        ws.column_dimensions["A"].width = 6

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name="Paysheet_Report.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500