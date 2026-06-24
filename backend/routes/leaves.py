from flask import Blueprint, request, jsonify
from models.database import db
from models.leave import LeaveRequest
from datetime import datetime, date
from models.employee import Employee
from openpyxl import Workbook
from flask import send_file
from io import BytesIO

from services.leave_balance_service import (
    update_all_employee_leave_balances  
)

leave_bp = Blueprint(
    "leave",
    __name__
)

@leave_bp.route("/", methods=["POST"])
def apply_leave():

    try:

        data = request.json

        leave = LeaveRequest(

            employee_id=data["employee_id"],

            employee_name=data["employee_name"],

            leave_type=data["leave_type"],

            from_date=datetime.strptime(
                data["from_date"],
                "%Y-%m-%d"
            ).date(),

            to_date=datetime.strptime(
                data["to_date"],
                "%Y-%m-%d"
            ).date(),

            total_days=data["total_days"],

            reporting_manager=data["reporting_manager"],

            handover_to=data["handover_to"],

            emergency_contact=data[
                "emergency_contact"
            ],

            reason=data["reason"]
        )

        db.session.add(leave)
        db.session.commit()

        return jsonify({
            "success": True,
            "message": "Leave Applied Successfully"
        })

    except Exception as e:

        db.session.rollback()

        return jsonify({
            "success": False,
            "error": str(e)
        }), 500
    
@leave_bp.route("/", methods=["GET"])
def get_leaves():

    leaves = LeaveRequest.query.order_by(
        LeaveRequest.id.desc()
    ).all()

    return jsonify([
        {
            "id": leave.id,

            "employee_id":
                leave.employee_id,

            "employee_name":
                leave.employee_name,

            "leave_type":
                leave.leave_type,

            "from_date":
                str(leave.from_date),

            "to_date":
                str(leave.to_date),

            "total_days":
                leave.total_days,

            "reporting_manager":
                leave.reporting_manager,

            "status":
                leave.status,

            "reason":
                leave.reason
        }

        for leave in leaves
    ])

@leave_bp.route(
"/approve/<int:leave_id>",
methods=["PUT"]
)
def approve_leave(leave_id):

    try:

        leave = LeaveRequest.query.get(leave_id)

        if not leave:
            return jsonify({
            "success": False,
            "error": "Leave not found"
            }), 404

    # Prevent double approval
        if leave.status == "Approved":
                return jsonify({
            "success": False,
            "error": "Leave already approved"
             }), 400

        print("Leave Employee ID:", leave.employee_id)
        

        employee = Employee.query.get(
        int(leave.employee_id)
        )
        print("Employee Found:", employee)

        if not employee:
          return jsonify({
            "success": False,
            "error": "Employee not found"
        }), 404

    # Update leave status
        leave.status = "Approved"

        leave_type = (
        leave.leave_type or ""
        ).strip().lower()

        leave_days = leave.total_days or 0

    # Deduct leave balance
        if leave_type == "sick leave":

           employee.sick_leave = max(
            0,
            (employee.sick_leave or 0) - leave_days
        )

        elif leave_type == "casual leave":

           employee.casual_leave = max(
            0,
            (employee.casual_leave or 0) - leave_days
          )

        elif leave_type == "earned leave":

           employee.earned_leave = max(
            0,
            (employee.earned_leave or 0) - leave_days
          )

        else:
          return jsonify({
            "success": False,
            "error": f"Invalid leave type: {leave.leave_type}"
        }), 400

        db.session.commit()

        return jsonify({
        "success": True,
        "message": "Leave Approved Successfully",
        "leave_balance": {

            "sick_leave":
                employee.sick_leave,

            "casual_leave":
                employee.casual_leave,

            "earned_leave":
                employee.earned_leave,

            "total_balance":
                (employee.sick_leave or 0) +
                (employee.casual_leave or 0) +
                (employee.earned_leave or 0)
            }
        }), 200

    except Exception as e:

      db.session.rollback()
      return jsonify({
        "success": False,
        "error": str(e)
    }), 500


@leave_bp.route(
    "/reject/<int:leave_id>",
    methods=["PUT"]
)
def reject_leave(leave_id):

    leave = LeaveRequest.query.get(
        leave_id
    )

    if not leave:

        return jsonify({
            "error": "Leave not found"
        }), 404

    leave.status = "Rejected"

    db.session.commit()

    return jsonify({
        "success": True,
        "message": "Leave Rejected"
    })


@leave_bp.route(
    "/cancel/<int:leave_id>",
    methods=["PUT"]
)
def cancel_leave(leave_id):

    try:

        leave = LeaveRequest.query.get(
            leave_id
        )

        if not leave:
            return jsonify({
                "success": False,
                "error": "Leave not found"
            }), 404

        if leave.status == "Cancelled":
            return jsonify({
                "success": False,
                "error": "Already cancelled"
            }), 400

        leave.status = "Cancelled"

        db.session.commit()

        return jsonify({
            "success": True,
            "message": "Leave Cancelled Successfully"
        })

    except Exception as e:

        db.session.rollback()

        return jsonify({
            "success": False,
            "error": str(e)
        }), 500
    
@leave_bp.route(
    "/update/<int:leave_id>",
    methods=["PUT"]
)
def update_leave(leave_id):

    try:

        leave = LeaveRequest.query.get(
            leave_id
        )

        if not leave:
            return jsonify({
                "success": False
            }), 404

        data = request.json

        leave.leave_type = data["leave_type"]

        leave.from_date = datetime.strptime(
            data["from_date"],
            "%Y-%m-%d"
        ).date()

        leave.to_date = datetime.strptime(
            data["to_date"],
            "%Y-%m-%d"
        ).date()

        leave.reason = data["reason"]

        leave.total_days = data["total_days"]

        db.session.commit()

        return jsonify({
            "success": True,
            "message": "Leave Updated"
        })

    except Exception as e:

        db.session.rollback()

        return jsonify({
            "success": False,
            "error": str(e)
        }), 500
    
@leave_bp.route(
    "/export-leave-report",
    methods=["GET"]
)
def export_leave_report():

    try:

        from io import BytesIO

        from openpyxl import Workbook

        from openpyxl.styles import (
            PatternFill,
            Font,
            Border,
            Side,
            Alignment
        )

        from openpyxl.utils import (
            get_column_letter
        )

        wb = Workbook()

        ws = wb.active

        ws.title = "Leave Working Update"

        today = date.today()

        if today.day < 25:

            if today.month == 1:

                start_date = date(
                    today.year - 1,
                    11,
                    25
            )

                end_date = date(
                    today.year - 1,
                    12,
                    24
            )        

            else:

                start_date = date(
                    today.year,
                    today.month - 2,
                    25
                )

                end_date = date(
                    today.year,
                    today.month - 1,
                    24
                )

        else:

            if today.month == 1:

                start_date = date(
                    today.year - 1,
                    12,
                    25
            )        

            else:

                start_date = date(
                    today.year,
                    today.month - 1,
                    25
                )

                end_date = date(
                    today.year,
                    today.month,
                    24
                )

        # =========================
        # STYLES
        # =========================

        blue_fill = PatternFill(
            "solid",
            fgColor="4F81BD"
        )

        green_fill = PatternFill(
            "solid",
            fgColor="00E5C3"
        )

        yellow_fill = PatternFill(
            "solid",
            fgColor="FFD966"
        )

        black_fill = PatternFill(
            "solid",
            fgColor="000000"
        )

        orange_fill = PatternFill(
            "solid",
            fgColor="F4B183"
        )

        white_font = Font(
            bold=True,
            color="FFFFFF"
        )

        bold_font = Font(
            bold=True
        )

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        # =========================
        # TITLE
        # =========================

        ws.merge_cells("A1:T1")

        ws["A1"] = "LEAVE WORKING UPDATE"

        ws["A1"].fill = blue_fill

        ws["A1"].font = Font(
            bold=True,
            size=16,
            color="FFFFFF"
        )

        ws["A1"].alignment = Alignment(
            horizontal="center"
        )

        # =========================
        # MONTH HEADER
        # =========================

        ws.merge_cells("A2:T2")

        ws["A2"] = (
            f"Leave Summary {today.strftime('%B %Y')}"
        )

        ws["A2"].fill = blue_fill

        ws["A2"].font = Font(
            bold=True,
            color="FFFFFF"
        )

        ws["A2"].alignment = Alignment(
            horizontal="center"
        )

        # =========================
        # LEAVE CYCLE
        # =========================

        ws.merge_cells("A3:T3")

        ws["A3"] = (
            f"Leave Cycle : "
            f"{start_date.strftime('%d-%b-%Y')} "
            f"to "
            f"{end_date.strftime('%d-%b-%Y')}"
        )

        ws["A3"].alignment = Alignment(
            horizontal="center"
        )

        # =========================
        # GROUP HEADERS
        # =========================

        ws.merge_cells("F4:H4")
        ws["F4"] = "Opening Balance"

        ws.merge_cells("I4:K4")
        ws["I4"] = "Monthly Credit"

        ws.merge_cells("L4:N4")
        ws["L4"] = "Leaves Deducted"

        ws.merge_cells("P4:R4")
        ws["P4"] = "Closing Balance"

        for cell in [
            "F4",
            "I4",
            "L4",
            "P4"
        ]:

            ws[cell].fill = black_fill
            ws[cell].font = white_font
            ws[cell].alignment = Alignment(
                horizontal="center"
            )

        # =========================
        # COLUMN HEADERS
        # =========================

        headers = [

            "Employee ID",
            "Employee Name",
            "Department",
            "Designation",
            "DOJ",

            "Opening CL",
            "Opening SL",
            "Opening EL",

            "CL Credit",
            "SL Credit",
            "EL Credit",

            "CL Taken",
            "SL Taken",
            "EL Taken",

            "Total Deducted",

            "Closing CL",
            "Closing SL",
            "Closing EL",

            "Total Balance",

            "Remarks"
        ]

        for col_num, header in enumerate(
            headers,
            start=1
        ):

            cell = ws.cell(
                row=5,
                column=col_num
            )

            cell.value = header
            cell.font = bold_font
            cell.border = thin_border
            cell.alignment = Alignment(
                horizontal="center"
            )

            if col_num <= 5:
                cell.fill = blue_fill

            elif 6 <= col_num <= 8:
                cell.fill = yellow_fill

            elif 9 <= col_num <= 11:
                cell.fill = green_fill

            elif 12 <= col_num <= 15:
                cell.fill = orange_fill

            elif 16 <= col_num <= 19:
                cell.fill = yellow_fill

        # =========================
        # EMPLOYEE DATA
        # =========================

        employees = Employee.query.order_by(
            Employee.first_name
        ).all()

        row = 6

        for employee in employees:

            cl_taken = 0
            sl_taken = 0
            el_taken = 0

            approved_leaves = LeaveRequest.query.filter(
                LeaveRequest.employee_id == str(employee.id),
                LeaveRequest.status == "Approved",
                LeaveRequest.from_date >= start_date,
                LeaveRequest.from_date <= end_date
            ).all()

            for leave in approved_leaves:

                leave_type = (
                    leave.leave_type or ""
                ).strip().lower()

                leave_days = (
                    leave.total_days or 0
                )

                if leave_type == "casual leave":
                    cl_taken += leave_days

                elif leave_type == "sick leave":
                    sl_taken += leave_days

                elif leave_type == "earned leave":
                    el_taken += leave_days

            current_cl = employee.casual_leave or 0
            current_sl = employee.sick_leave or 0
            current_el = employee.earned_leave or 0

            credit_cl = 1.5
            credit_sl = 1.5
            credit_el = 2

            opening_cl = current_cl + credit_cl
            opening_sl = current_sl + credit_sl
            opening_el = current_el + credit_el

            closing_cl = opening_cl - cl_taken
            closing_sl = opening_sl - sl_taken
            closing_el = opening_el - el_taken

            total_deducted = (
                cl_taken +
                sl_taken +
                el_taken
            )

            total_balance = (
                closing_cl +
                closing_sl +
                closing_el
            )

            ws.cell(row=row, column=1, value=employee.employee_id)
            ws.cell(row=row, column=2, value=f"{employee.first_name} {employee.last_name}")
            ws.cell(row=row, column=3, value=employee.department)
            ws.cell(row=row, column=4, value=employee.designation)
            ws.cell(row=row, column=5, value=str(employee.joining_date))

            ws.cell(row=row, column=6, value=opening_cl)
            ws.cell(row=row, column=7, value=opening_sl)
            ws.cell(row=row, column=8, value=opening_el)

            ws.cell(row=row, column=9, value=credit_cl)
            ws.cell(row=row, column=10, value=credit_sl)
            ws.cell(row=row, column=11, value=credit_el)

            ws.cell(row=row, column=12, value=cl_taken)
            ws.cell(row=row, column=13, value=sl_taken)
            ws.cell(row=row, column=14, value=el_taken)

            ws.cell(row=row, column=15, value=total_deducted)

            ws.cell(row=row, column=16, value=closing_cl)
            ws.cell(row=row, column=17, value=closing_sl)
            ws.cell(row=row, column=18, value=closing_el)

            ws.cell(row=row, column=19, value=total_balance)

            ws.cell(row=row, column=20, value="")

            for col in range(1, 21):

                cell = ws.cell(
                    row=row,
                    column=col
                )

                cell.border = thin_border

                # Employee ID, DOJ, and all numeric columns -> center
                if col in [
                    1,
                    5,
                    6, 7, 8, 9, 10, 11,
                    12, 13, 14, 15,
                    16, 17, 18, 19
                ]:

                    cell.alignment = Alignment(
                        horizontal="center",
                        vertical="center"
                    )

                # Employee Name, Department, Designation, Remarks -> left
                else:

                    cell.alignment = Alignment(
                        horizontal="left",
                        vertical="center"
                    )

            row += 1

        # =========================
        # AUTO WIDTH
        # =========================

        for column_cells in ws.columns:

            length = max(
                len(str(cell.value))
                if cell.value
                else 0
                for cell in column_cells
            )

            ws.column_dimensions[
                get_column_letter(
                    column_cells[0].column
                )
            ].width = length + 5

        # Keep Employee ID column compact regardless of auto width
        ws.column_dimensions["A"].width = 12

        # =========================
        # FILTER
        # =========================

        ws.auto_filter.ref = f"A5:T{row}"

        # =========================
        # DOWNLOAD
        # =========================

        excel_file = BytesIO()

        wb.save(excel_file)

        excel_file.seek(0)

        return send_file(
            excel_file,
            as_attachment=True,
            download_name="Leave_Working_Update.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:

        return jsonify({
            "success": False,
            "error": str(e)
        }), 500

@leave_bp.route(
    "/credit-monthly-leaves",
    methods=["POST"]
)
def credit_monthly_leaves():

    try:

        result = (
            update_all_employee_leave_balances()
        )

        return jsonify(result)

    except Exception as e:

        db.session.rollback()

        return jsonify({
            "success": False,
            "error": str(e)
        }), 500